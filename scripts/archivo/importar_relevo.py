"""
Importación de la ESTRATEGIA DE RELEVO GENERACIONAL a la base de datos.

Lee los dos Excel de partida y los carga en el modelo relacional:
  - Seguimiento_proyecto.xlsx  -> plan, ámbitos, actuaciones, seguimientos (bitácora)
  - KPI.xlsx                   -> indicadores y sus valores por año

Es re-ejecutable: borra y recrea el plan RELEVO antes de cargar, de modo
que puedes volver a lanzarlo si actualizas los Excel.

Uso:
    python scripts/importar_relevo.py
Coloca los dos Excel en la carpeta  datos/  del proyecto.
"""
import re
import sys
from pathlib import Path

import openpyxl

# Permitir importar src/db.py
RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ / "src"))
import db  # noqa: E402

DATOS = RAIZ / "datos"
F_SEGUIMIENTO = DATOS / "Seguimiento_proyecto.xlsx"
F_KPI = DATOS / "KPI.xlsx"

# Patrón de cabecera de ámbito: "5.1. Marco Normativo", "5.2 Valorización..."
RE_AMBITO = re.compile(r"^(\d+\.\d+)\.?\s+(.*)$")


# --------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------
def limpiar(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_presupuesto(v):
    """Devuelve (importe_float_o_None, nota_o_None)."""
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip()
    if not s:
        return None, None
    # ¿es un número escrito como texto?
    limpio = s.replace("€", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(limpio), None
    except ValueError:
        return None, s  # ej. 'Sin dotación', 'N/D'


def obtener_o_crear_responsable(con, nombre):
    nombre = nombre.strip()
    fila = con.execute(
        "SELECT id FROM responsables WHERE nombre = ?", (nombre,)
    ).fetchone()
    if fila:
        return fila["id"]
    cur = con.execute(
        "INSERT INTO responsables (nombre, organizacion) VALUES (?, ?)",
        (nombre, nombre),
    )
    return cur.lastrowid


def vincular_responsables(con, actuacion_id, texto_responsable):
    """'GV - DDFF - HAZI' -> tres responsables vinculados."""
    if not texto_responsable:
        return
    partes = re.split(r"\s*[-/]\s*", texto_responsable)
    for p in partes:
        p = p.strip()
        if p:
            rid = obtener_o_crear_responsable(con, p)
            con.execute(
                "INSERT OR IGNORE INTO actuacion_responsables "
                "(actuacion_id, responsable_id) VALUES (?, ?)",
                (actuacion_id, rid),
            )


# --------------------------------------------------------------------------
# Carga del plan + ámbitos + actuaciones + seguimientos
# --------------------------------------------------------------------------
def importar_seguimiento(con):
    wb = openpyxl.load_workbook(F_SEGUIMIENTO, data_only=True)
    ws = wb["Seguimiento proyecto"]

    # Crear el plan
    cur = con.execute(
        """INSERT INTO planes (codigo, nombre_es, departamento,
                               periodo_inicio, periodo_fin, objetivo_macro_es, estado)
           VALUES (?,?,?,?,?,?,?)""",
        (
            "RELEVO",
            "Estrategia de Relevo Generacional",
            "Alimentación, Desarrollo Rural, Agricultura y Pesca",
            2025,
            2028,
            None,
            "Activo",
        ),
    )
    plan_id = cur.lastrowid

    ambito_id = None
    orden_ambito = 0
    orden_act = 0
    n_act = 0

    filas = list(ws.iter_rows(values_only=True))
    for fila in filas:
        col_a = limpiar(fila[0]) if len(fila) > 0 else None
        if not col_a:
            continue

        # ¿cabecera de ámbito?  ej. "5.1. Marco Normativo"
        m = RE_AMBITO.match(col_a)
        if m:
            codigo, nombre = m.group(1), m.group(2).strip()
            orden_ambito += 1
            orden_act = 0
            cur = con.execute(
                "INSERT INTO ambitos (plan_id, codigo, nombre_es, orden) VALUES (?,?,?,?)",
                (plan_id, codigo, nombre, orden_ambito),
            )
            ambito_id = cur.lastrowid
            continue

        # saltar fila de cabeceras de columnas
        if col_a == "Actuación":
            continue

        # impacto global (última fila) -> objetivo macro del plan
        if col_a.startswith("✅") or "Impacto global" in col_a:
            texto = col_a.replace("✅", "").replace("Impacto global esperado:", "").strip()
            con.execute("UPDATE planes SET objetivo_macro_es = ? WHERE id = ?",
                        (texto, plan_id))
            continue

        # si no hay ámbito todavía, no es una actuación
        if ambito_id is None:
            continue

        # --- es una actuación ---
        nombre_act = col_a
        objetivo = limpiar(fila[1]) if len(fila) > 1 else None
        presup_raw = fila[2] if len(fila) > 2 else None
        fecha_prev = limpiar(fila[4]) if len(fila) > 4 else None
        responsable = limpiar(fila[5]) if len(fila) > 5 else None
        detalle_dic = limpiar(fila[7]) if len(fila) > 7 else None
        detalle_may = limpiar(fila[8]) if len(fila) > 8 else None

        importe, nota = parse_presupuesto(presup_raw)

        # estado inicial: si hay algún detalle de seguimiento -> 'En curso'
        # (es una estimación de partida; se ajusta luego en la aplicación)
        estado = "En curso" if (detalle_dic or detalle_may) else "Previsto"

        orden_act += 1
        cur = con.execute(
            """INSERT INTO actuaciones
               (ambito_id, nombre_es, objetivo_impacto_es, presupuesto,
                presupuesto_nota, fecha_inicio_prevista, estado, orden)
               VALUES (?,?,?,?,?,?,?,?)""",
            (ambito_id, nombre_act, objetivo, importe, nota,
             fecha_prev, estado, orden_act),
        )
        act_id = cur.lastrowid
        n_act += 1

        vincular_responsables(con, act_id, responsable)

        # Bitácora: un registro por cada corte con detalle
        if detalle_dic:
            con.execute(
                """INSERT INTO seguimientos
                   (actuacion_id, fecha_corte, etiqueta_corte, estado, detalle_es)
                   VALUES (?,?,?,?,?)""",
                (act_id, "2025-12-01", "DIC 2025", estado, detalle_dic),
            )
        if detalle_may:
            con.execute(
                """INSERT INTO seguimientos
                   (actuacion_id, fecha_corte, etiqueta_corte, estado, detalle_es)
                   VALUES (?,?,?,?,?)""",
                (act_id, "2026-05-01", "MAY 2026", estado, detalle_may),
            )

    con.commit()
    print(f"  Plan RELEVO: {orden_ambito} ámbitos, {n_act} actuaciones cargadas.")
    return plan_id


# --------------------------------------------------------------------------
# Carga de indicadores (KPI)
# --------------------------------------------------------------------------
RE_NUM = re.compile(r"^\s*(\d+)\s*[\.\)]?\s*(.*)$", re.DOTALL)

# Normalización de categorías
CATEGORIAS = {
    "RESULTADO": "Resultado / Impacto",
    "EJECUCIÓN": "Ejecución",
    "EJECUCION": "Ejecución",
    "APOYO": "Apoyo y seguimiento",
}


def normaliza_categoria(texto):
    if not texto:
        return None
    t = texto.upper()
    for clave, valor in CATEGORIAS.items():
        if clave in t:
            return valor
    return texto.strip()


def importar_kpi(con, plan_id):
    wb = openpyxl.load_workbook(F_KPI, data_only=True)
    ws = wb["seguimiento KPIs"]

    # columnas (0-indexadas): B=categoria(1), C=indicador(2), D=def(3),
    # E=meta(4), F=desarrollo(5), G..J = 2025..2028 (6..9)
    anios = {6: "2025", 7: "2026", 8: "2027", 9: "2028"}

    categoria_actual = None
    orden = 0
    n_ind = 0

    for fila in ws.iter_rows(values_only=True):
        cat = limpiar(fila[1]) if len(fila) > 1 else None
        if cat:
            categoria_actual = normaliza_categoria(cat)

        nombre = limpiar(fila[2]) if len(fila) > 2 else None
        if not nombre:
            continue
        # la fila de cabecera lleva 'INDICADOR'
        if nombre.upper() == "INDICADOR":
            continue

        # separar número del nombre  ("1. \nTasa de incorporaciones..." )
        m = RE_NUM.match(nombre.replace("\n", " ").strip())
        if m:
            numero = int(m.group(1))
            nombre_limpio = m.group(2).strip()
        else:
            numero, nombre_limpio = None, nombre

        definicion = limpiar(fila[3]) if len(fila) > 3 else None
        meta = limpiar(fila[4]) if len(fila) > 4 else None
        desarrollo = limpiar(fila[5]) if len(fila) > 5 else None

        orden += 1
        cur = con.execute(
            """INSERT INTO indicadores
               (plan_id, categoria, numero, nombre_es, definicion_es,
                meta_es, desarrollo_es, orden)
               VALUES (?,?,?,?,?,?,?,?)""",
            (plan_id, categoria_actual, numero, nombre_limpio,
             definicion, meta, desarrollo, orden),
        )
        ind_id = cur.lastrowid
        n_ind += 1

        # valores por año
        for col, periodo in anios.items():
            v = fila[col] if len(fila) > col else None
            if v is None:
                continue
            if isinstance(v, (int, float)):
                con.execute(
                    "INSERT INTO indicador_valores (indicador_id, periodo, valor) VALUES (?,?,?)",
                    (ind_id, periodo, float(v)),
                )
            else:
                texto = str(v).strip()
                if texto:
                    con.execute(
                        "INSERT INTO indicador_valores "
                        "(indicador_id, periodo, valor_texto_es) VALUES (?,?,?)",
                        (ind_id, periodo, texto),
                    )

    con.commit()
    print(f"  Indicadores (KPI): {n_ind} cargados.")


# --------------------------------------------------------------------------
def main():
    if not F_SEGUIMIENTO.exists() or not F_KPI.exists():
        print("ERROR: coloca 'Seguimiento_proyecto.xlsx' y 'KPI.xlsx' en la carpeta datos/")
        sys.exit(1)

    db.inicializar_db()
    con = db.conectar()

    # Re-ejecutable: borra el plan RELEVO previo (cascada limpia lo dependiente)
    con.execute("DELETE FROM planes WHERE codigo = 'RELEVO'")
    con.commit()

    print("Importando Estrategia de Relevo Generacional...")
    plan_id = importar_seguimiento(con)
    importar_kpi(con, plan_id)

    con.close()
    print("Importación completada.")


if __name__ == "__main__":
    main()
