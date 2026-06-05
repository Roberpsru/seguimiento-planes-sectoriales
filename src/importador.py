"""
Lógica de importación/exportación de Planes Sectoriales en formato Excel
estandarizado (`datos/Plan_Sectorial_<NOMBRE>.xlsx`).

Esta capa la usan dos sitios:
  - `scripts/importar_plan.py` — envoltorio CLI fino.
  - `vistas/5_Administracion.py` — página de carga/descarga desde la UI.

API pública:
  - cargar_plan_desde_excel(origen, reemplazar=False, dry_run=False) -> dict
  - exportar_plan_a_excel(plan_id) -> bytes
"""
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db


# --------------------------------------------------------------------------
# Constantes compartidas con scripts/importar_plan.py
# --------------------------------------------------------------------------
ESTADOS_VALIDOS = {"Previsto", "En curso", "Ejecutado"}

CATEGORIAS = {
    "RESULTADO": "Resultado / Impacto",
    "EJECUCION": "Ejecución",
    "APOYO":     "Apoyo y seguimiento",
}

MESES_ABREV_ES = [
    "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGO", "SEP", "OCT", "NOV", "DIC",
]

COLOR_VERDE = "1F5F3A"

# Columnas y anchos por hoja, para que el Excel exportado sea visualmente
# idéntico a las plantillas de partida.
_HOJAS = {
    "Plan": {
        "titulo_cols": 8,
        "cabeceras": [
            "Código *", "Nombre castellano *", "Nombre euskera",
            "Periodo inicio *", "Periodo fin *",
            "Organización promotora *",
            "Descripción castellano", "Descripción euskera",
        ],
        "anchos": {1: 12, 2: 38, 4: 13, 6: 36, 7: 60, 8: 32},
    },
    "Ámbitos": {
        "titulo_cols": 4,
        "cabeceras": [
            "Código del ámbito *", "Nombre castellano *",
            "Nombre euskera", "Orden *",
        ],
        "anchos": {1: 22, 2: 40, 4: 12},
    },
    "Responsables": {
        "titulo_cols": 6,
        "cabeceras": [
            "Código *\nKodea *",
            "Nombre castellano *\nGaztelaniazko izena *",
            "Nombre euskera\nEuskarazko izena",
            "Organización castellano\nGaztelaniazko erakundea",
            "Organización euskera\nEuskarazko erakundea",
            "Email",
        ],
        "anchos": {1: 14, 2: 40, 3: 40, 4: 32, 5: 32, 6: 28},
    },
    "Actuaciones": {
        "titulo_cols": 11,
        "cabeceras": [
            "Código actuación *", "Ámbito (código) *",
            "Nombre castellano *", "Nombre euskera",
            "Descripción castellano", "Descripción euskera",
            "Presupuesto total (€) *",
            "Fecha inicio prevista", "Fecha fin prevista",
            "Responsable (código)", "Estado inicial *",
        ],
        "anchos": {1: 12, 2: 14, 3: 32, 4: 22, 5: 55, 6: 22,
                   7: 16, 8: 14, 10: 16, 11: 14},
    },
    "Indicadores": {
        "titulo_cols": 12,
        "cabeceras": [
            "Código KPI *", "Nombre castellano *", "Nombre euskera",
            "Tipo *", "Unidad *",
            "Definición castellano *", "Definición euskera",
            "Valor línea base", "Año línea base",
            "Meta valor *", "Meta año *", "Orden *",
        ],
        "anchos": {1: 12, 2: 32, 3: 28, 4: 14, 6: 55, 7: 28,
                   8: 13, 11: 12, 12: 10},
    },
    "Valores_indicadores": {
        "titulo_cols": 5,
        "cabeceras": [
            "Código KPI *\nKPI kodea *",
            "Año *\nUrtea *",
            "Valor *\nBalioa *",
            "Observaciones castellano\nGaztelaniazko oharrak",
            "Observaciones euskera\nEuskarazko oharrak",
        ],
        "anchos": {1: 14, 2: 10, 3: 16, 4: 60, 5: 60},
    },
    "Seguimientos": {
        "titulo_cols": 9,
        "cabeceras": [
            "Código actuación *\nJarduketaren kodea *",
            "Fecha de corte *\nMozketa-data *",
            "Etiqueta\nEtiketa",
            "Estado *\nEgoera *",
            "Presupuesto ejecutado\nExekutatutako aurrekontua",
            "Fecha inicio real\nBenetako hasiera-data",
            "Fecha fin real\nBenetako amaiera-data",
            "Observaciones castellano\nGaztelaniazko oharrak",
            "Observaciones euskera\nEuskarazko oharrak",
        ],
        "anchos": {1: 16, 2: 14, 3: 12, 4: 14, 5: 18, 6: 16, 7: 16, 8: 50, 9: 50},
    },
}


# --------------------------------------------------------------------------
# Helpers de parsing
# --------------------------------------------------------------------------
def _norm(s):
    if s is None:
        return ""
    n = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in n if not unicodedata.combining(c)).strip().lower()


def _hoja(wb, nombre_esperado):
    objetivo = _norm(nombre_esperado)
    for sh in wb.sheetnames:
        if _norm(sh) == objetivo:
            return wb[sh]
    raise ValueError(
        f"No se encuentra la hoja '{nombre_esperado}'. "
        f"Disponibles: {wb.sheetnames}"
    )


def _limpiar(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _entero(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _decimal(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace(".", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _fecha_iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Fecha no reconocida: {v!r}")


def _filas_datos(ws):
    for fila in ws.iter_rows(min_row=4, values_only=True):
        if any(c not in (None, "") for c in fila):
            yield fila


# --------------------------------------------------------------------------
# Lectura POR NOMBRE DE CABECERA (más robusta que por posición)
#
# Cada hoja tiene su fila de cabecera (normalmente la fila 3). Las cabeceras
# del Excel rediseñado son bilingües en DOS líneas dentro de la misma celda:
# castellano arriba, euskera debajo ("Nombre castellano *\nGaztelaniazko izena *").
# Para casar la columna sólo nos interesa la parte CASTELLANA (antes del \n).
# --------------------------------------------------------------------------
def _norm_cab(valor):
    """Normaliza una cabecera para casarla por nombre.

    - Toma sólo la parte castellana (lo anterior al primer salto de línea).
    - Sustituye '_' por espacio (acepta variantes tipo 'Nombre_eu').
    - Quita asteriscos y acentos, pasa a minúsculas y colapsa espacios.

    Ej.: 'Nombre castellano *\\nGaztelaniazko izena *' -> 'nombre castellano'
         'Año *'                                       -> 'ano'
         'Nombre_eu'                                   -> 'nombre eu'
    """
    if valor is None:
        return ""
    texto = str(valor).split("\n")[0]
    texto = texto.replace("_", " ").replace("*", " ")
    texto = _norm(texto)                 # NFKD + sin acentos + minúsculas + strip
    return " ".join(texto.split())       # colapsa espacios internos


def _mapa_cabeceras(ws):
    """Devuelve ({cabecera_normalizada: índice_col_0based}, fila_cabecera).

    La fila de cabecera es la primera fila no vacía a partir de la fila 3
    (título en la 1, fila 2 en blanco, cabecera en la 3 en las plantillas).
    """
    fila_cab = None
    for r in range(3, 8):
        celdas = [c.value for c in ws[r]]
        if any(v not in (None, "") for v in celdas):
            fila_cab = r
            break
    if fila_cab is None:
        fila_cab = 3
    mapa = {}
    for idx, cell in enumerate(ws[fila_cab]):
        clave = _norm_cab(cell.value)
        if clave and clave not in mapa:
            mapa[clave] = idx
    return mapa, fila_cab


def _filas_datos_cab(ws, fila_cab):
    """Itera las filas de datos (tuplas de valores) tras la fila de cabecera."""
    for fila in ws.iter_rows(min_row=fila_cab + 1, values_only=True):
        if any(c not in (None, "") for c in fila):
            yield fila


def _val(fila, mapa, candidatos, fn=_limpiar):
    """Valor de la primera cabecera de `candidatos` presente en la fila.

    `candidatos` es una lista de claves YA normalizadas (con _norm_cab).
    Si ninguna está en el mapa de cabeceras, devuelve None (la columna no
    existe en este Excel -> el campo no se rellena). `fn` transforma el valor
    bruto (_limpiar por defecto; _entero / _decimal / _fecha_iso / identidad).
    """
    for clave in candidatos:
        idx = mapa.get(clave)
        if idx is not None and idx < len(fila):
            return fn(fila[idx])
    return None


def _parse_categoria(texto):
    if not texto:
        return None
    t = _norm(texto).upper()
    for clave, valor in CATEGORIAS.items():
        if clave in t:
            return valor
    return str(texto).strip()


def _parse_numero_kpi(codigo):
    if not codigo:
        return None
    digitos = ""
    for ch in str(codigo):
        if ch.isdigit():
            digitos += ch
        elif digitos:
            break
    return int(digitos) if digitos else None


def _etiqueta_desde_fecha(fecha_iso):
    try:
        d = datetime.strptime(fecha_iso, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None
    return f"{MESES_ABREV_ES[d.month - 1]} {d.year}"


_RE_META_ANIO = re.compile(r"\s+en\s+(20\d{2})\s*$")


# --------------------------------------------------------------------------
# Lectura del Excel (acepta ruta o file-like / bytes)
# --------------------------------------------------------------------------
def _abrir_wb(origen):
    """Acepta ruta (str/Path), bytes o file-like."""
    if isinstance(origen, (str, Path)):
        return openpyxl.load_workbook(origen, data_only=True)
    if isinstance(origen, bytes):
        return openpyxl.load_workbook(BytesIO(origen), data_only=True)
    # file-like (BytesIO, UploadedFile…): debe soportar .read()
    return openpyxl.load_workbook(origen, data_only=True)


def _leer_excel(origen):
    wb = _abrir_wb(origen)

    # ---- Plan (fila única) ----
    ws = _hoja(wb, "Plan")
    mapa, fcab = _mapa_cabeceras(ws)
    fila = next(_filas_datos_cab(ws, fcab), None)
    if not fila:
        raise ValueError("La hoja 'Plan' está vacía.")
    plan = {
        "codigo":         _val(fila, mapa, ["codigo"]),
        "nombre_es":      _val(fila, mapa, ["nombre castellano", "nombre"]),
        "nombre_eu":      _val(fila, mapa, ["nombre euskera", "nombre eu"]),
        "periodo_inicio": _val(fila, mapa, ["periodo inicio"], _entero),
        "periodo_fin":    _val(fila, mapa, ["periodo fin"], _entero),
        "departamento":   _val(fila, mapa, ["organizacion promotora", "organizacion"]),
        "descripcion_es": _val(fila, mapa, ["descripcion castellano", "descripcion"]),
        "descripcion_eu": _val(fila, mapa, ["descripcion euskera", "descripcion eu"]),
    }

    # ---- Ámbitos ----
    ws = _hoja(wb, "Ámbitos")
    mapa, fcab = _mapa_cabeceras(ws)
    ambitos = []
    for fila in _filas_datos_cab(ws, fcab):
        amb = {
            "codigo":    _val(fila, mapa, ["codigo del ambito", "codigo"]),
            "nombre_es": _val(fila, mapa, ["nombre castellano", "nombre"]),
            "nombre_eu": _val(fila, mapa, ["nombre euskera", "nombre eu"]),
            "orden":     _val(fila, mapa, ["orden"], _entero),
        }
        if amb["codigo"] and amb["nombre_es"]:
            ambitos.append(amb)

    # ---- Responsables (NUEVO: nombre_eu / organizacion_eu) ----
    ws = _hoja(wb, "Responsables")
    mapa, fcab = _mapa_cabeceras(ws)
    responsables = []
    for fila in _filas_datos_cab(ws, fcab):
        r = {
            "codigo":          _val(fila, mapa, ["codigo"]),
            "nombre":          _val(fila, mapa, ["nombre castellano", "nombre"]),
            "nombre_eu":       _val(fila, mapa, ["nombre euskera", "nombre eu"]),
            "organizacion":    _val(fila, mapa, ["organizacion castellano", "organizacion"]),
            "organizacion_eu": _val(fila, mapa, ["organizacion euskera", "organizacion eu"]),
            "email":           _val(fila, mapa, ["email"]),
        }
        if r["codigo"] and r["nombre"]:
            responsables.append(r)

    # ---- Actuaciones ----
    ws = _hoja(wb, "Actuaciones")
    mapa, fcab = _mapa_cabeceras(ws)
    actuaciones = []
    for fila in _filas_datos_cab(ws, fcab):
        presup_raw = _val(
            fila, mapa,
            ["presupuesto total (€)", "presupuesto total", "presupuesto"],
            lambda x: x,
        )
        a = {
            "codigo":          _val(fila, mapa, ["codigo actuacion", "codigo"]),
            "ambito_codigo":   _val(fila, mapa, ["ambito (codigo)", "ambito codigo", "ambito"]),
            "nombre_es":       _val(fila, mapa, ["nombre castellano", "nombre"]),
            "nombre_eu":       _val(fila, mapa, ["nombre euskera", "nombre eu"]),
            "descripcion_es":  _val(fila, mapa, ["descripcion castellano", "descripcion"]),
            "descripcion_eu":  _val(fila, mapa, ["descripcion euskera", "descripcion eu"]),
            "presupuesto":     _decimal(presup_raw),
            "presupuesto_raw": presup_raw,
            "fecha_inicio":    _val(fila, mapa, ["fecha inicio prevista", "fecha inicio"], _fecha_iso),
            "fecha_fin":       _val(fila, mapa, ["fecha fin prevista", "fecha fin"], _fecha_iso),
            "responsable_cod": _val(fila, mapa, ["responsable (codigo)", "responsable codigo", "responsable"]),
            "estado":          _val(fila, mapa, ["estado inicial", "estado"]),
        }
        if a["codigo"] and a["nombre_es"]:
            actuaciones.append(a)

    # ---- Indicadores ----
    ws = _hoja(wb, "Indicadores")
    mapa, fcab = _mapa_cabeceras(ws)
    indicadores = []
    for fila in _filas_datos_cab(ws, fcab):
        ind = {
            "codigo":           _val(fila, mapa, ["codigo kpi", "codigo"]),
            "nombre_es":        _val(fila, mapa, ["nombre castellano", "nombre"]),
            "nombre_eu":        _val(fila, mapa, ["nombre euskera", "nombre eu"]),
            "tipo":             _val(fila, mapa, ["tipo"]),
            "unidad":           _val(fila, mapa, ["unidad"]),
            "definicion_es":    _val(fila, mapa, ["definicion castellano", "definicion"]),
            "definicion_eu":    _val(fila, mapa, ["definicion euskera", "definicion eu"]),
            "valor_linea_base": _val(fila, mapa, ["valor linea base"], _decimal),
            "anio_linea_base":  _val(fila, mapa, ["ano linea base"], _entero),
            "meta_valor":       _val(fila, mapa, ["meta valor"], _decimal),
            "meta_anio":        _val(fila, mapa, ["meta ano"], _entero),
            "orden":            _val(fila, mapa, ["orden"], _entero),
        }
        if ind["codigo"] and ind["nombre_es"]:
            indicadores.append(ind)

    # ---- Valores_indicadores (NUEVO: observaciones_eu -> nota_eu) ----
    ws = _hoja(wb, "Valores_indicadores")
    mapa, fcab = _mapa_cabeceras(ws)
    valores = []
    for fila in _filas_datos_cab(ws, fcab):
        v = {
            "kpi_codigo":       _val(fila, mapa, ["codigo kpi", "codigo"]),
            "anio":             _val(fila, mapa, ["ano"], _entero),
            "valor":            _val(fila, mapa, ["valor"], _decimal),
            "observaciones":    _val(fila, mapa, ["observaciones castellano", "observaciones"]),
            "observaciones_eu": _val(fila, mapa, ["observaciones euskera", "observaciones eu"]),
        }
        if v["kpi_codigo"]:
            valores.append(v)

    # ---- Seguimientos (NUEVO: observaciones_eu -> detalle_eu) ----
    ws = _hoja(wb, "Seguimientos")
    mapa, fcab = _mapa_cabeceras(ws)
    seguimientos = []
    for fila in _filas_datos_cab(ws, fcab):
        cod = _val(fila, mapa, ["codigo actuacion", "codigo"])
        if not cod or cod.startswith("("):
            continue
        s = {
            "actuacion_cod":     cod,
            "fecha_corte":       _val(fila, mapa, ["fecha de corte", "fecha corte"], _fecha_iso),
            "etiqueta":          _val(fila, mapa, ["etiqueta"]),
            "estado":            _val(fila, mapa, ["estado"]),
            "presupuesto_ejec":  _val(fila, mapa, ["presupuesto ejecutado"], _decimal),
            "fecha_inicio_real": _val(fila, mapa, ["fecha inicio real"], _fecha_iso),
            "fecha_fin_real":    _val(fila, mapa, ["fecha fin real"], _fecha_iso),
            "observaciones":     _val(fila, mapa, ["observaciones castellano", "observaciones"]),
            "observaciones_eu":  _val(fila, mapa, ["observaciones euskera", "observaciones eu"]),
        }
        seguimientos.append(s)

    return {
        "plan":         plan,
        "ambitos":      ambitos,
        "responsables": responsables,
        "actuaciones":  actuaciones,
        "indicadores":  indicadores,
        "valores":      valores,
        "seguimientos": seguimientos,
    }


# --------------------------------------------------------------------------
# Validaciones (recopilan TODAS las incidencias antes de devolverlas)
# --------------------------------------------------------------------------
def _validar(datos):
    err = []
    plan = datos["plan"]

    if not plan["codigo"]:
        err.append("Plan: falta el código.")
    if not plan["nombre_es"]:
        err.append("Plan: falta el nombre en castellano.")

    def _unicos(bloque, items):
        vistos = {}
        for offset, it in enumerate(items):
            fila_real = offset + 4
            cod = it["codigo"]
            if not cod:
                err.append(f"{bloque} (fila {fila_real}): falta el código.")
                continue
            if cod in vistos:
                err.append(
                    f"{bloque}: código duplicado '{cod}' "
                    f"(filas {vistos[cod]} y {fila_real})."
                )
            else:
                vistos[cod] = fila_real

    _unicos("Ámbitos",      datos["ambitos"])
    _unicos("Responsables", datos["responsables"])
    _unicos("Actuaciones",  datos["actuaciones"])
    _unicos("Indicadores",  datos["indicadores"])

    cod_ambitos      = {a["codigo"] for a in datos["ambitos"]}
    cod_responsables = {r["codigo"] for r in datos["responsables"]}
    cod_actuaciones  = {a["codigo"] for a in datos["actuaciones"]}
    cod_indicadores  = {i["codigo"] for i in datos["indicadores"]}

    for offset, a in enumerate(datos["actuaciones"]):
        fila_real = offset + 4
        if a["ambito_codigo"] not in cod_ambitos:
            err.append(
                f"Actuaciones (fila {fila_real}, '{a['codigo']}'): "
                f"ámbito '{a['ambito_codigo']}' no existe."
            )
        if a["responsable_cod"] and a["responsable_cod"] not in cod_responsables:
            err.append(
                f"Actuaciones (fila {fila_real}, '{a['codigo']}'): "
                f"responsable '{a['responsable_cod']}' no existe."
            )
        if a["estado"] and a["estado"] not in ESTADOS_VALIDOS:
            err.append(
                f"Actuaciones (fila {fila_real}, '{a['codigo']}'): "
                f"estado '{a['estado']}' no es válido "
                f"(usa Previsto / En curso / Ejecutado)."
            )
        if a["fecha_inicio"] and a["fecha_fin"] and a["fecha_inicio"] > a["fecha_fin"]:
            err.append(
                f"Actuaciones (fila {fila_real}, '{a['codigo']}'): "
                f"fecha inicio ({a['fecha_inicio']}) posterior a "
                f"fecha fin ({a['fecha_fin']})."
            )

    for offset, v in enumerate(datos["valores"]):
        fila_real = offset + 4
        if v["kpi_codigo"] not in cod_indicadores:
            err.append(
                f"Valores_indicadores (fila {fila_real}): KPI "
                f"'{v['kpi_codigo']}' no existe."
            )
        if v["anio"] is None:
            err.append(
                f"Valores_indicadores (fila {fila_real}, "
                f"'{v['kpi_codigo']}'): falta el año."
            )

    for offset, s in enumerate(datos["seguimientos"]):
        fila_real = offset + 4
        if s["actuacion_cod"] not in cod_actuaciones:
            err.append(
                f"Seguimientos (fila {fila_real}): actuación "
                f"'{s['actuacion_cod']}' no existe."
            )
        if s["estado"] and s["estado"] not in ESTADOS_VALIDOS:
            err.append(
                f"Seguimientos (fila {fila_real}, '{s['actuacion_cod']}'): "
                f"estado '{s['estado']}' no válido."
            )
        if (s["fecha_inicio_real"] and s["fecha_fin_real"]
                and s["fecha_inicio_real"] > s["fecha_fin_real"]):
            err.append(
                f"Seguimientos (fila {fila_real}, '{s['actuacion_cod']}'): "
                f"fecha inicio real posterior a fecha fin real."
            )

    return err


# --------------------------------------------------------------------------
# Inserción (toda en una transacción)
# --------------------------------------------------------------------------
def _insertar(con, datos):
    plan = datos["plan"]

    plan_id = db.insertar_devolver_id(
        con, "planes",
        ["codigo", "nombre_es", "nombre_eu",
         "descripcion_es", "descripcion_eu",
         "departamento", "periodo_inicio", "periodo_fin",
         "objetivo_macro_es", "estado"],
        [
            plan["codigo"], plan["nombre_es"], plan["nombre_eu"],
            plan["descripcion_es"], plan["descripcion_eu"],
            plan["departamento"], plan["periodo_inicio"], plan["periodo_fin"],
            plan["descripcion_es"], "Activo",
        ],
    )

    map_ambitos = {}
    for a in datos["ambitos"]:
        map_ambitos[a["codigo"]] = db.insertar_devolver_id(
            con, "ambitos",
            ["plan_id", "codigo", "nombre_es", "nombre_eu", "orden"],
            [plan_id, a["codigo"], a["nombre_es"], a["nombre_eu"], a["orden"] or 0],
        )

    map_responsables = {}
    for r in datos["responsables"]:
        org = r["organizacion"] or r["codigo"]
        fila = con.execute(
            "SELECT id FROM responsables "
            " WHERE nombre = ? AND COALESCE(organizacion,'') = COALESCE(?,'')",
            (r["nombre"], org),
        ).fetchone()
        if fila:
            # Ya existe (los responsables NO se borran al reemplazar un plan,
            # son globales). Rellenamos su euskera si está vacío, SIN sobrescribir
            # un valor ya presente (COALESCE deja el actual si no es NULL). Así un
            # responsable creado por un import antiguo recibe sus traducciones al
            # volver a subir el Excel con las columnas _eu.
            con.execute(
                "UPDATE responsables "
                "   SET nombre_eu       = COALESCE(nombre_eu, ?), "
                "       organizacion_eu = COALESCE(organizacion_eu, ?) "
                " WHERE id = ?",
                (r.get("nombre_eu"), r.get("organizacion_eu"), fila["id"]),
            )
            map_responsables[r["codigo"]] = fila["id"]
        else:
            map_responsables[r["codigo"]] = db.insertar_devolver_id(
                con, "responsables",
                ["nombre", "nombre_eu", "organizacion", "organizacion_eu", "email"],
                [r["nombre"], r.get("nombre_eu"), org, r.get("organizacion_eu"),
                 r["email"]],
            )

    map_actuaciones = {}
    orden_por_ambito = {}
    for a in datos["actuaciones"]:
        ambito_id = map_ambitos[a["ambito_codigo"]]
        orden_por_ambito[a["ambito_codigo"]] = (
            orden_por_ambito.get(a["ambito_codigo"], 0) + 1
        )
        presup = a["presupuesto"]
        nota = None
        if presup is None and a["presupuesto_raw"] is not None:
            s = str(a["presupuesto_raw"]).strip()
            if s:
                nota = s
        act_id = db.insertar_devolver_id(
            con, "actuaciones",
            ["ambito_id", "codigo", "nombre_es", "nombre_eu",
             "objetivo_impacto_es", "objetivo_impacto_eu",
             "presupuesto", "presupuesto_nota",
             "fecha_inicio_prevista", "fecha_fin_prevista",
             "estado", "orden"],
            [
                ambito_id, a["codigo"], a["nombre_es"], a["nombre_eu"],
                a["descripcion_es"], a["descripcion_eu"],
                presup, nota,
                a["fecha_inicio"], a["fecha_fin"],
                a["estado"] or "Previsto",
                orden_por_ambito[a["ambito_codigo"]],
            ],
        )
        map_actuaciones[a["codigo"]] = act_id

        if a["responsable_cod"]:
            # "No duplicar" agnóstico de motor: INSERT OR IGNORE (sqlite) /
            # ON CONFLICT DO NOTHING (postgres).
            if db.MOTOR_BD == "postgres":
                con.execute(
                    "INSERT INTO actuacion_responsables "
                    "(actuacion_id, responsable_id) VALUES (?,?) "
                    "ON CONFLICT DO NOTHING",
                    (act_id, map_responsables[a["responsable_cod"]]),
                )
            else:
                con.execute(
                    "INSERT OR IGNORE INTO actuacion_responsables "
                    "(actuacion_id, responsable_id) VALUES (?,?)",
                    (act_id, map_responsables[a["responsable_cod"]]),
                )

    map_indicadores = {}
    for ind in datos["indicadores"]:
        meta_es = None
        if ind["meta_valor"] is not None and ind["meta_anio"]:
            meta_es = f"{ind['meta_valor']:g} en {ind['meta_anio']}"
        elif ind["meta_valor"] is not None:
            meta_es = f"{ind['meta_valor']:g}"
        map_indicadores[ind["codigo"]] = db.insertar_devolver_id(
            con, "indicadores",
            ["plan_id", "categoria", "numero", "nombre_es", "nombre_eu",
             "definicion_es", "definicion_eu",
             "meta_es", "meta_valor", "unidad", "orden"],
            [
                plan_id,
                _parse_categoria(ind["tipo"]),
                _parse_numero_kpi(ind["codigo"]),
                ind["nombre_es"], ind["nombre_eu"],
                ind["definicion_es"], ind["definicion_eu"],
                meta_es, ind["meta_valor"], ind["unidad"],
                ind["orden"] or 0,
            ],
        )

    for v in datos["valores"]:
        ind_id = map_indicadores[v["kpi_codigo"]]
        con.execute(
            """INSERT INTO indicador_valores
                   (indicador_id, periodo, valor, nota_es, nota_eu)
               VALUES (?,?,?,?,?)""",
            (ind_id, str(v["anio"]), v["valor"],
             v["observaciones"], v.get("observaciones_eu")),
        )

    for s in datos["seguimientos"]:
        act_id = map_actuaciones[s["actuacion_cod"]]
        etiqueta = s["etiqueta"] or _etiqueta_desde_fecha(s["fecha_corte"])
        con.execute(
            """INSERT INTO seguimientos
                   (actuacion_id, fecha_corte, etiqueta_corte,
                    estado, detalle_es, detalle_eu, presupuesto_ejecutado)
               VALUES (?,?,?,?,?,?,?)""",
            (act_id, s["fecha_corte"], etiqueta,
             s["estado"], s["observaciones"], s.get("observaciones_eu"),
             s["presupuesto_ejec"]),
        )

    return plan_id


def _impacto_actual(con, codigo):
    """Cuenta lo que se borraría si se reemplaza el plan con ese código."""
    actuaciones = con.execute(
        """SELECT COUNT(*) AS n FROM actuaciones a
             JOIN ambitos am ON a.ambito_id = am.id
             JOIN planes  p  ON am.plan_id = p.id
            WHERE p.codigo = ?""",
        (codigo,),
    ).fetchone()["n"]
    seguimientos = con.execute(
        """SELECT COUNT(*) AS n FROM seguimientos s
             JOIN actuaciones a ON s.actuacion_id = a.id
             JOIN ambitos am ON a.ambito_id = am.id
             JOIN planes p ON am.plan_id = p.id
            WHERE p.codigo = ?""",
        (codigo,),
    ).fetchone()["n"]
    indicadores = con.execute(
        """SELECT COUNT(*) AS n FROM indicadores i
             JOIN planes p ON i.plan_id = p.id
            WHERE p.codigo = ?""",
        (codigo,),
    ).fetchone()["n"]
    valores = con.execute(
        """SELECT COUNT(*) AS n FROM indicador_valores iv
             JOIN indicadores i ON iv.indicador_id = i.id
             JOIN planes p ON i.plan_id = p.id
            WHERE p.codigo = ?""",
        (codigo,),
    ).fetchone()["n"]
    return {
        "actuaciones":  actuaciones,
        "seguimientos": seguimientos,
        "indicadores":  indicadores,
        "valores":      valores,
    }


# --------------------------------------------------------------------------
# API pública — CARGA
# --------------------------------------------------------------------------
def cargar_plan_desde_excel(origen, reemplazar=False, dry_run=False):
    """
    Lee un Excel, lo valida y opcionalmente lo carga en la BD.

    Parámetros:
      origen      ruta a .xlsx (str/Path), bytes o file-like (BytesIO, etc.)
      reemplazar  si True, borra el plan existente (mismo código) antes de cargar.
      dry_run     si True, no inserta nada en la BD (solo valida y reporta).

    Devuelve un dict con la forma descrita en la docstring del módulo.
    """
    resumen_vacio = {
        "ambitos": 0, "actuaciones": 0, "responsables": 0,
        "indicadores": 0, "valores": 0, "seguimientos": 0,
    }
    base = {
        "ok": False,
        "plan_codigo": None,
        "plan_nombre": None,
        "plan_existia_antes": False,
        "accion": "error",
        "resumen": dict(resumen_vacio),
        "afectado_si_reemplaza": None,
        "incidencias": [],
        "mensaje": "",
    }

    # 1) Leer el Excel
    try:
        datos = _leer_excel(origen)
    except Exception as e:
        base["mensaje"] = f"No se ha podido leer el Excel: {e}"
        base["incidencias"] = [base["mensaje"]]
        return base

    plan = datos["plan"]
    base["plan_codigo"] = plan["codigo"] or ""
    base["plan_nombre"] = plan["nombre_es"] or ""
    base["resumen"] = {
        "ambitos":      len(datos["ambitos"]),
        "actuaciones":  len(datos["actuaciones"]),
        "responsables": len(datos["responsables"]),
        "indicadores":  len(datos["indicadores"]),
        "valores":      len(datos["valores"]),
        "seguimientos": len(datos["seguimientos"]),
    }

    # 2) Validar
    incidencias = _validar(datos)
    if incidencias:
        base["incidencias"] = incidencias
        base["mensaje"] = f"{len(incidencias)} incidencia(s) en el Excel."
        return base

    # 3) Comprobar existencia y calcular impacto si procede
    con = db.conectar()
    try:
        fila = con.execute(
            "SELECT id FROM planes WHERE codigo = ?", (plan["codigo"],)
        ).fetchone()
        base["plan_existia_antes"] = fila is not None
        if base["plan_existia_antes"]:
            base["afectado_si_reemplaza"] = _impacto_actual(con, plan["codigo"])

        # 4) Modo simulación
        if dry_run:
            base["ok"] = True
            base["accion"] = "preview"
            base["mensaje"] = (
                f"Excel válido. Plan '{plan['codigo']}'."
            )
            return base

        # 5) Carga real
        if base["plan_existia_antes"] and not reemplazar:
            base["mensaje"] = (
                f"El plan '{plan['codigo']}' ya existe en la BD. "
                f"Usa --reemplazar para sobrescribirlo."
            )
            base["incidencias"] = [base["mensaje"]]
            return base

        try:
            if base["plan_existia_antes"] and reemplazar:
                con.execute(
                    "DELETE FROM planes WHERE codigo = ?", (plan["codigo"],)
                )
            _insertar(con, datos)
            con.commit()
        except Exception as e:
            con.rollback()
            base["mensaje"] = f"Fallo al insertar; transacción revertida: {e}"
            base["incidencias"] = [base["mensaje"]]
            return base

        base["ok"] = True
        base["accion"] = "reemplazado" if base["plan_existia_antes"] else "creado"
        base["mensaje"] = (
            f"Plan {plan['codigo']} cargado: "
            f"{base['resumen']['ambitos']} ámbitos, "
            f"{base['resumen']['actuaciones']} actuaciones, "
            f"{base['resumen']['indicadores']} indicadores, "
            f"{base['resumen']['valores']} valores, "
            f"{base['resumen']['seguimientos']} seguimientos."
        )
        return base
    finally:
        con.close()


# --------------------------------------------------------------------------
# API pública — EXPORTACIÓN
# --------------------------------------------------------------------------
def _crear_hoja(wb, nombre, titulo_es):
    """Crea una hoja con el formato estándar (título fila 1, cabeceras fila 3)."""
    spec = _HOJAS[nombre]
    ws = wb.create_sheet(nombre)

    # Fila 1: título grande en verde, fondo blanco, merged.
    ws.cell(row=1, column=1, value=titulo_es)
    ws.cell(row=1, column=1).font = Font(
        name="Calibri", size=18, bold=True, color=COLOR_VERDE,
    )
    ncols = spec["titulo_cols"]
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    # Fila 3: cabecera (verde con texto blanco bold).
    fill = PatternFill(start_color=COLOR_VERDE, end_color=COLOR_VERDE, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for i, etiqueta in enumerate(spec["cabeceras"], start=1):
        c = ws.cell(row=3, column=i, value=etiqueta)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Anchos de columna.
    for col_idx, ancho in spec["anchos"].items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    return ws


_TIPO_INVERSO = {
    "Resultado / Impacto": "Resultado",
    "Ejecución":           "Ejecución",
    "Apoyo y seguimiento": "Apoyo",
}


def _codigo_responsable(r, ya_usados):
    """Inventa un código único para un responsable, basado en su organización
    o, en su defecto, en un sufijo R1, R2..."""
    org = (r["organizacion"] or "").strip()
    if org and len(org) <= 12 and org not in ya_usados:
        ya_usados.add(org)
        return org
    base = (org or r["nombre"] or "RESP")[:8].upper().replace(" ", "_")
    i = 1
    cand = base
    while cand in ya_usados:
        i += 1
        cand = f"{base}_{i}"
    ya_usados.add(cand)
    return cand


def _periodo_a_anio(periodo):
    """'2025' → 2025; cualquier otra cosa → el texto original."""
    try:
        return int(periodo)
    except (TypeError, ValueError):
        return periodo


def _fecha_iso_a_date(fecha_iso):
    """'YYYY-MM-DD' → date; None / inválido → None."""
    if not fecha_iso:
        return None
    try:
        return datetime.strptime(fecha_iso[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def exportar_plan_a_excel(plan_id):
    """
    Genera un Excel con el estado actual del plan, en el mismo formato que
    las plantillas `datos/Plan_Sectorial_<NOMBRE>.xlsx`. Devuelve los bytes
    listos para `st.download_button`.
    """
    con = db.conectar()
    try:
        plan = con.execute(
            "SELECT * FROM planes WHERE id = ?", (plan_id,)
        ).fetchone()
        if not plan:
            raise ValueError(f"Plan id={plan_id} no encontrado.")
        plan = dict(plan)

        ambitos = [dict(r) for r in con.execute(
            "SELECT * FROM ambitos WHERE plan_id = ? ORDER BY orden, id",
            (plan_id,),
        ).fetchall()]
        ambito_id_to_codigo = {a["id"]: a["codigo"] for a in ambitos}

        actuaciones = [dict(r) for r in con.execute(
            """SELECT ac.* FROM actuaciones ac
                 JOIN ambitos am ON ac.ambito_id = am.id
                WHERE am.plan_id = ?
                ORDER BY am.orden, ac.orden, ac.id""",
            (plan_id,),
        ).fetchall()]

        # Responsable principal por actuación (el de menor id en la N:M).
        primario = {}
        for r in con.execute(
            """SELECT ar.actuacion_id, MIN(ar.responsable_id) AS resp_id
                 FROM actuacion_responsables ar
                 JOIN actuaciones a ON ar.actuacion_id = a.id
                 JOIN ambitos am   ON a.ambito_id     = am.id
                WHERE am.plan_id = ?
                GROUP BY ar.actuacion_id""",
            (plan_id,),
        ).fetchall():
            primario[r["actuacion_id"]] = r["resp_id"]

        # Responsables únicos que aparecen en este plan.
        resp_ids_plan = sorted(set(primario.values()))
        responsables = []
        if resp_ids_plan:
            qs = ",".join("?" for _ in resp_ids_plan)
            for r in con.execute(
                f"SELECT * FROM responsables WHERE id IN ({qs}) ORDER BY id",
                resp_ids_plan,
            ).fetchall():
                responsables.append(dict(r))

        indicadores = [dict(r) for r in con.execute(
            "SELECT * FROM indicadores WHERE plan_id = ? ORDER BY orden, numero, id",
            (plan_id,),
        ).fetchall()]

        valores = [dict(r) for r in con.execute(
            """SELECT iv.* FROM indicador_valores iv
                 JOIN indicadores i ON iv.indicador_id = i.id
                WHERE i.plan_id = ?
                ORDER BY i.orden, iv.periodo, iv.id""",
            (plan_id,),
        ).fetchall()]

        seguimientos = [dict(r) for r in con.execute(
            """SELECT s.* , ac.codigo AS act_codigo FROM seguimientos s
                 JOIN actuaciones ac ON s.actuacion_id = ac.id
                 JOIN ambitos am    ON ac.ambito_id   = am.id
                WHERE am.plan_id = ?
                ORDER BY s.fecha_corte DESC, s.id DESC""",
            (plan_id,),
        ).fetchall()]
    finally:
        con.close()

    # Códigos para los responsables: se calculan una vez y se reutilizan en
    # la hoja Responsables y en la columna Responsable (código) de Actuaciones.
    ya_usados = set()
    cod_resp_por_id = {}
    for r in responsables:
        cod = _codigo_responsable(r, ya_usados)
        cod_resp_por_id[r["id"]] = cod

    # Códigos de actuación e indicador (los que ya están en BD; si fueran
    # null, se inventan a partir del orden).
    cod_act_por_id = {}
    for i, ac in enumerate(actuaciones, start=1):
        cod_act_por_id[ac["id"]] = ac.get("codigo") or f"A{i}"

    cod_ind_por_id = {}
    for ind in indicadores:
        cod = ind.get("codigo")
        if not cod:
            num = ind.get("numero")
            cod = f"KPI-{num:02d}" if num else f"KPI-{ind['id']}"
        cod_ind_por_id[ind["id"]] = cod

    # ---- Construcción del workbook ----
    wb = openpyxl.Workbook()
    # La hoja por defecto la usamos como "Instrucciones".
    ws_instr = wb.active
    ws_instr.title = "Instrucciones"
    ws_instr.column_dimensions["A"].width = 2
    ws_instr.column_dimensions["B"].width = 100
    ws_instr.cell(row=2, column=2, value=plan["nombre_es"]).font = Font(
        size=16, bold=True, color=COLOR_VERDE,
    )
    nota = (
        f"Este Excel se generó desde la aplicación el "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')} con el estado "
        f"actual del Plan «{plan['nombre_es']}». Para subirlo de vuelta, "
        f"usar la página de Administración con la opción Reemplazar."
    )
    ws_instr.cell(row=4, column=2, value=nota).alignment = Alignment(wrap_text=True)
    ws_instr.row_dimensions[4].height = 60

    # ---- Plan ----
    ws = _crear_hoja(wb, "Plan", plan["nombre_es"] or plan["codigo"] or "Plan")
    ws.cell(row=4, column=1, value=plan["codigo"])
    ws.cell(row=4, column=2, value=plan["nombre_es"])
    ws.cell(row=4, column=3, value=plan["nombre_eu"])
    ws.cell(row=4, column=4, value=plan["periodo_inicio"])
    ws.cell(row=4, column=5, value=plan["periodo_fin"])
    ws.cell(row=4, column=6, value=plan["departamento"])
    ws.cell(row=4, column=7, value=plan["descripcion_es"])
    ws.cell(row=4, column=8, value=plan["descripcion_eu"])

    # ---- Ámbitos ----
    ws = _crear_hoja(wb, "Ámbitos", "Ejes de intervención")
    for i, a in enumerate(ambitos, start=4):
        ws.cell(row=i, column=1, value=a["codigo"])
        ws.cell(row=i, column=2, value=a["nombre_es"])
        ws.cell(row=i, column=3, value=a["nombre_eu"])
        ws.cell(row=i, column=4, value=a["orden"])

    # ---- Responsables ----
    ws = _crear_hoja(wb, "Responsables", "Responsables")
    for i, r in enumerate(responsables, start=4):
        ws.cell(row=i, column=1, value=cod_resp_por_id[r["id"]])
        ws.cell(row=i, column=2, value=r["nombre"])
        ws.cell(row=i, column=3, value=r.get("nombre_eu"))
        ws.cell(row=i, column=4, value=r["organizacion"])
        ws.cell(row=i, column=5, value=r.get("organizacion_eu"))
        ws.cell(row=i, column=6, value=r["email"])

    # ---- Actuaciones ----
    ws = _crear_hoja(wb, "Actuaciones", f"Actuaciones — {len(actuaciones)} actuaciones")
    for i, a in enumerate(actuaciones, start=4):
        presup = a["presupuesto"]
        if presup is None and a["presupuesto_nota"]:
            presup = a["presupuesto_nota"]
        ws.cell(row=i, column=1, value=a["codigo"])
        ws.cell(row=i, column=2, value=ambito_id_to_codigo.get(a["ambito_id"]))
        ws.cell(row=i, column=3, value=a["nombre_es"])
        ws.cell(row=i, column=4, value=a["nombre_eu"])
        ws.cell(row=i, column=5, value=a["objetivo_impacto_es"])
        ws.cell(row=i, column=6, value=a["objetivo_impacto_eu"])
        ws.cell(row=i, column=7, value=presup)
        ws.cell(row=i, column=8, value=_fecha_iso_a_date(a["fecha_inicio_prevista"])
                or a["fecha_inicio_prevista"])
        ws.cell(row=i, column=9, value=_fecha_iso_a_date(a["fecha_fin_prevista"])
                or a["fecha_fin_prevista"])
        resp_id = primario.get(a["id"])
        ws.cell(row=i, column=10, value=cod_resp_por_id.get(resp_id) if resp_id else None)
        ws.cell(row=i, column=11, value=a["estado"])
        # Formato fecha para las columnas 8 y 9 si el valor es date.
        for col in (8, 9):
            v = ws.cell(row=i, column=col).value
            if isinstance(v, (date, datetime)):
                ws.cell(row=i, column=col).number_format = "DD/MM/YYYY"

    # ---- Indicadores ----
    ws = _crear_hoja(wb, "Indicadores", f"Indicadores (KPI) — {len(indicadores)} indicadores")
    for i, ind in enumerate(indicadores, start=4):
        # Recupera Meta año si está embebido en meta_es ("712.5 en 2030").
        meta_anio = None
        if ind["meta_es"]:
            m = _RE_META_ANIO.search(ind["meta_es"])
            if m:
                meta_anio = int(m.group(1))

        tipo = _TIPO_INVERSO.get(ind["categoria"], ind["categoria"])
        ws.cell(row=i, column=1, value=cod_ind_por_id[ind["id"]])
        ws.cell(row=i, column=2, value=ind["nombre_es"])
        ws.cell(row=i, column=3, value=ind["nombre_eu"])
        ws.cell(row=i, column=4, value=tipo)
        ws.cell(row=i, column=5, value=ind["unidad"])
        ws.cell(row=i, column=6, value=ind["definicion_es"])
        ws.cell(row=i, column=7, value=ind["definicion_eu"])
        # Valor / Año línea base: no se guardan como tales en BD.
        ws.cell(row=i, column=8, value=None)
        ws.cell(row=i, column=9, value=None)
        ws.cell(row=i, column=10, value=ind["meta_valor"])
        ws.cell(row=i, column=11, value=meta_anio)
        ws.cell(row=i, column=12, value=ind["orden"])

    # ---- Valores_indicadores ----
    ws = _crear_hoja(wb, "Valores_indicadores", "Valores anuales de indicadores")
    # Mapa indicador_id → código KPI (texto en el Excel).
    cod_kpi_por_indid = cod_ind_por_id
    fila_excel = 4
    for v in valores:
        kpi_cod = cod_kpi_por_indid.get(v["indicador_id"])
        if not kpi_cod:
            continue
        ws.cell(row=fila_excel, column=1, value=kpi_cod)
        ws.cell(row=fila_excel, column=2, value=_periodo_a_anio(v["periodo"]))
        ws.cell(row=fila_excel, column=3, value=v["valor"])
        ws.cell(row=fila_excel, column=4, value=v["nota_es"])
        ws.cell(row=fila_excel, column=5, value=v.get("nota_eu"))
        fila_excel += 1

    # ---- Seguimientos ----
    ws = _crear_hoja(wb, "Seguimientos", "Historial de seguimiento")
    fila_excel = 4
    for s in seguimientos:
        ws.cell(row=fila_excel, column=1, value=s["act_codigo"])
        f_corte = _fecha_iso_a_date(s["fecha_corte"])
        ws.cell(row=fila_excel, column=2,
                value=f_corte or s["fecha_corte"])
        if isinstance(f_corte, (date, datetime)):
            ws.cell(row=fila_excel, column=2).number_format = "DD/MM/YYYY"
        ws.cell(row=fila_excel, column=3, value=s["etiqueta_corte"])
        ws.cell(row=fila_excel, column=4, value=s["estado"])
        ws.cell(row=fila_excel, column=5, value=s["presupuesto_ejecutado"])
        # Fecha inicio/fin real no se guardan en el esquema actual.
        ws.cell(row=fila_excel, column=6, value=None)
        ws.cell(row=fila_excel, column=7, value=None)
        ws.cell(row=fila_excel, column=8, value=s["detalle_es"])
        ws.cell(row=fila_excel, column=9, value=s.get("detalle_eu"))
        fila_excel += 1

    # Volcado a bytes.
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
